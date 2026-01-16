"""
Data export utilities for CSV and Excel
"""

import csv
from io import StringIO, BytesIO
from typing import List, Dict, Any
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from datetime import datetime


def export_to_csv(data: List[Dict[str, Any]], headers: List[str]) -> str:
    """
    Export data to CSV format
    
    Args:
        data: List of dictionaries containing data
        headers: List of column headers
        
    Returns:
        str: CSV content
    """
    output = StringIO()
    writer = csv.DictWriter(output, fieldnames=headers)
    
    writer.writeheader()
    writer.writerows(data)
    
    return output.getvalue()


def export_to_excel(data: List[Dict[str, Any]], headers: List[str], sheet_name: str = "Data") -> bytes:
    """
    Export data to Excel format
    
    Args:
        data: List of dictionaries containing data
        headers: List of column headers
        sheet_name: Name of the Excel sheet
        
    Returns:
        bytes: Excel file content
    """
    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name
    
    # Style for headers
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")
    
    # Write headers
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
    
    # Write data
    for row_num, row_data in enumerate(data, 2):
        for col_num, header in enumerate(headers, 1):
            value = row_data.get(header, "")
            
            # Format datetime objects
            if isinstance(value, datetime):
                value = value.strftime("%Y-%m-%d %H:%M:%S")
            
            ws.cell(row=row_num, column=col_num, value=value)
    
    # Auto-adjust column widths
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width
    
    # Save to BytesIO
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    
    return output.getvalue()


def prepare_sales_export_data(sales: List[Any]) -> tuple[List[Dict], List[str]]:
    """Prepare sales data for export"""
    headers = [
        "Transaction ID",
        "Customer Name",
        "Product ID",
        "Quantity",
        "Unit Price",
        "Total Amount",
        "Payment Method",
        "Created At",
        "Notes"
    ]
    
    data = []
    for sale in sales:
        data.append({
            "Transaction ID": sale.transaction_id,
            "Customer Name": sale.customer_name,
            "Product ID": sale.product_id,
            "Quantity": sale.quantity,
            "Unit Price": float(sale.unit_price),
            "Total Amount": float(sale.total_amount),
            "Payment Method": sale.payment_method,
            "Created At": sale.created_at,
            "Notes": sale.notes or ""
        })
    
    return data, headers


def prepare_inventory_export_data(items: List[Any]) -> tuple[List[Dict], List[str]]:
    """Prepare inventory data for export"""
    headers = [
        "SKU",
        "Name",
        "Category",
        "Quantity",
        "Reorder Level",
        "Unit Price",
        "Supplier",
        "Location",
        "Status",
        "Created At"
    ]
    
    data = []
    for item in items:
        data.append({
            "SKU": item.sku,
            "Name": item.name,
            "Category": item.category or "",
            "Quantity": item.quantity,
            "Reorder Level": item.reorder_level,
            "Unit Price": float(item.unit_price),
            "Supplier": item.supplier or "",
            "Location": item.location or "",
            "Status": item.status,
            "Created At": item.created_at
        })
    
    return data, headers
